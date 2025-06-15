#app_accounting views.py
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import UpdateView
from django.urls import reverse_lazy
from django.utils.translation import gettext_lazy as _
from django.contrib import admin
from django.contrib.auth import get_user_model
from django.views.generic import DetailView
# from .models import Profile
from .forms import UserProfileForm # Вам нужно будет создать эту форму
from rest_framework.views import APIView
from rest_framework.response import Response
from app_productions.serializers import ProductionBatchSerializer
from logger import logger

import datetime
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required 
from django.utils import timezone
from django.db.models import Sum, F, Value, CharField
from django.db.models.functions import Concat
from django.shortcuts import redirect, get_object_or_404
from django.contrib.auth.decorators import login_required # или permission_required
from django.contrib import messages
from app_users.models import User 
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.urls import reverse_lazy
from django.utils.translation import gettext_lazy as _
from datetime import date
from app_accounting.models import UserSalary 
from app_productions.models import ProcessStage, WorkLog
from django.shortcuts import get_object_or_404, Http404 


def dashboard_callback(request, context):
    """
    Обновляет переданный контекст, добавляя виджеты и данные для шаблона,
    и возвращает ИЗМЕНЕННЫЙ контекст.
    """
    widgets = [
        {
            "title": _("Мой профиль"),
            "icon": "account_circle",
            "actions": [
                {
                    "title": _("Редактировать профиль"),
                    "link": reverse_lazy("accounts:profile"),
                    "icon": "edit",
                }
            ],
            "description": _("Просмотр и редактирование личных данных."),
            "permission": lambda req: req.user.is_authenticated,
        },
    ]

    context.update({
        "dashboard_widgets": widgets,
        "total_users": User.objects.count(),
        "latest_user": User.objects.order_by('-date_joined').first(),
    })

    return context

# --- Ваш UserProfileView ---
class UserProfileView(LoginRequiredMixin, UpdateView):
    model = User
    form_class = UserProfileForm
    template_name = 'admin/profile/profile_form.html' 
    success_url = reverse_lazy('accounts:profile')

    def get_object(self, queryset=None):
        return self.request.user 

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(admin.site.each_context(self.request))
        context['title'] = 'Мой профиль'
        return context

# --- Ваш UserProfileView ---
class UserProfileView(LoginRequiredMixin, UpdateView):
    model = User
    form_class = UserProfileForm
    template_name = 'admin/profile/profile_form.html' 
    success_url = reverse_lazy('accounts:profile')

    def get_object(self, queryset=None):
        return self.request.user 

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(admin.site.each_context(self.request))
        context['title'] = 'Мой профиль'
        return context
   
class UserSalaryReportView(LoginRequiredMixin, DetailView):
    model = UserSalary
    template_name = "admin/profile/profile_form.html"
    context_object_name = "usersalary" # Это имя будет использоваться в шаблоне для current_period_object

    def get_object(self, queryset=None):
        user = self.request.user
        latest_active_period = UserSalary.objects.filter(
            user=user,
            is_paid=False
        ).order_by('-period_start_datetime').first()
        return latest_active_period

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # self.object здесь будет равен результату get_object(), 
        # и также доступен в шаблоне как {{ usersalary }} из-за context_object_name
        current_period_object = self.object 

        context['title'] = _("Мой текущий расчетный период") # Изменил заголовок для ясности
        context['current_profile_user'] = self.request.user

        if current_period_object:
            context['no_salary_data_found'] = False
            # context['user_role'] = _("Сотрудник") # Замените на реальную логику получения роли, если нужно
            # Пример:
            if hasattr(self.request.user, 'employee_profile') and self.request.user.employee_profile.role:
                 context['user_role'] = self.request.user.employee_profile.get_role_display()
            else:
                 context['user_role'] = _("Роль не указана")



            total_items_aggregation = current_period_object.work_log.aggregate(
                total_processed=Sum('quantity_processed', default=0)
            )
            context['total_items_processed'] = total_items_aggregation['total_processed']

            # --- ИСПРАВЛЕННЫЙ SELECT_RELATED ---
            context['related_work_logs'] = current_period_object.work_log.select_related(
                'stage',                            
                'stage__batch_product',             # ProcessStage -> BatchProductLink (это одна выборка)
                'stage__batch_product__batch',      # BatchProductLink -> Batch (это вторая выборка, зависит от первой)
                'stage__batch_product__product'     # BatchProductLink -> Product (это третья выборка, зависит от первой)
            ).order_by('-log_time').all()

            if not current_period_object.is_paid:
                context['mark_paid_url'] = reverse_lazy('accounts:mark_salary_paid', kwargs={'pk': current_period_object.pk})
        else:
            context['no_salary_data_found'] = True
            # Явно передаем None, чтобы шаблон не падал при обращении к usersalary.что-то_там
            context['usersalary'] = None
            context['related_work_logs'] = [] 
            context['total_items_processed'] = 0


        # История ВСЕХ периодов для пользователя
        # Убедимся, что также делаем select_related для paid_by, если он используется в шаблоне истории
        context['user_salary_history'] = UserSalary.objects.filter(
            user=self.request.user
        ).select_related('paid_by').order_by('-period_start_datetime')

        return context


@login_required
def mark_salary_as_paid(request, pk):
    user_salary_to_close = get_object_or_404(UserSalary, pk=pk)

    can_mark_paid = request.user.is_superuser or request.user.has_perm('app_accounting.can_mark_salary_paid') or request.user == user_salary_to_close.user
    # Замените 'app_accounting.can_mark_salary_paid' на ваше реальное право доступа, если оно есть

    if not can_mark_paid:
        messages.error(request, _("У вас нет прав для выполнения этого действия."))
        return redirect(reverse_lazy('accounts:profile')) # Убедитесь, что 'accounts:profile' - это правильный URL

    if request.method == 'POST':
        if hasattr(user_salary_to_close, 'mark_as_paid_and_prepare_next'):
            success, _ = user_salary_to_close.mark_as_paid_and_prepare_next(request.user) # _ для new_salary_period, если он не используется дальше
            if success:
                messages.success(request, _("Расчетный период успешно закрыт. Если это был последний активный, новый будет создан при следующей записи работы."))
            else:
                messages.info(request, _("Этот расчетный период уже был закрыт ранее или произошла ошибка."))
        else:
            # Базовая логика, если метода mark_as_paid_and_prepare_next нет
            if not user_salary_to_close.is_paid:
                user_salary_to_close.is_paid = True
                user_salary_to_close.paid_at = timezone.now()
                user_salary_to_close.paid_by = request.user
                user_salary_to_close.save()
                messages.success(request, _("Расчетный период успешно закрыт."))
            else:
                messages.info(request, _("Этот расчетный период уже был закрыт ранее."))
    else:
        messages.warning(request, _("Некорректный метод запроса. Используйте POST."))

    return redirect(reverse_lazy('accounts:profile'))


def monthly_user_production_report(request):
    now = timezone.now()
    current_year = now.year
    current_month = now.month

    first_day_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    if current_month == 12:
        first_day_of_next_month = first_day_of_month.replace(year=current_year + 1, month=1)
    else:
        first_day_of_next_month = first_day_of_month.replace(month=current_month + 1)

    report_queryset = ProcessStage.objects.filter(
        end_date__gte=first_day_of_month,
        end_date__lt=first_day_of_next_month,
        status='COMPLETED',
        assigned_user__isnull=False
    ).select_related('assigned_user')

    report_data = report_queryset.values(
        'assigned_user',
        'stage_type'
    ).annotate(
        user_full_name=Concat(
            F('assigned_user__last_name'), Value(' '), F('assigned_user__first_name'),
            output_field=CharField()
        ),
        username=F('assigned_user__username'),
        total_completed=Sum('quantity_completed', default=0) # Добавил default=0
    ).order_by(
        'user_full_name',
        'stage_type'
    ).values(
        'user_full_name',
        'username',
        'stage_type',
        'total_completed'
    )

    if request.GET.get('export') == 'xlsx':
        wb = Workbook()
        ws = wb.active
        
        month_names = {
            1: _("Январь"), 2: _("Февраль"), 3: _("Март"), 4: _("Апрель"),
            5: _("Май"), 6: _("Июнь"), 7: _("Июль"), 8: _("Август"),
            9: _("Сентябрь"), 10: _("Октябрь"), 11: _("Ноябрь"), 12: _("Декабрь"),
        }
        current_month_name = month_names.get(current_month, str(current_month))

        ws.title = _("Отчет {} {}").format(current_month_name, current_year)
        ws.append([_("Отчет по производству за {} {}").format(current_month_name, current_year)])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4) # Объединение ячеек для заголовка
        ws['A1'].alignment = Alignment(horizontal='center') # Выравнивание заголовка по центру
        ws.append([]) 

        headers = [_('Пользователь (ФИО)'), _('Псевдоним'), _('Тип Этапа'), _('Выполнено (шт.)')]
        ws.append(headers)
        for cell in ws[ws.max_row]: # Жирный шрифт для заголовков таблицы
            cell.font = Font(bold=True)


        stage_display_names = dict(ProcessStage.STAGE_CHOICES)

        for item in report_data:
            stage_name = stage_display_names.get(item['stage_type'], item['stage_type'])
            ws.append([
                item['user_full_name'],
                item['username'],
                stage_name,
                item['total_completed']
            ])

        for col_idx, _ in enumerate(headers, 1):
             column_letter = get_column_letter(col_idx)
             ws.column_dimensions[column_letter].auto_size = True # Автоподбор ширины

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f"production_report_{current_year}_{current_month:02d}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    context = {
        'report_data': list(report_data),
        'current_month_name': current_month_name, # Передаем имя месяца
        'current_year': current_year,
        'stage_choices_dict': dict(ProcessStage.STAGE_CHOICES),
        'page_title': _("Отчет по производству за {} {}").format(current_month_name, current_year)
    }
    return render(request, 'reports/monthly_user_production_report.html', context)

# Закомментированный код APIView оставляю как есть, так как он не затрагивался
# class ProductionBatchAPIView(APIView):
#     def post(self, request):
#         serializer = ProductionBatchSerializer(data=request.data)
#         if not serializer.is_valid():
#             logger.error(f"Error serializers: {serializer.error}")
#             return Response(serializer.error, status=400)

# class ProductionBatchAPIView(APIView):
#     def post(self, request):
#         serializer = ProductionBatchSerializer(data=request.data)
#         if not serializer.is_valid():
#             logger.error(f"Error serializers: {serializer.error}")
#             return Response(serializer.error, status=400)


